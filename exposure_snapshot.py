import openpyxl
from datetime import datetime

# ================= CONFIG =================
EXCEL_PATH = "Macro_Data.xlsx"
SHEET_NAME = "Exposure"

# ================= MAIN =================
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

# Read the LAST row (latest exposure)
last_row = ws.max_row

equity = ws[f"B{last_row}"].value
notional = ws[f"C{last_row}"].value

if equity is None or notional is None:
    raise RuntimeError("Equity or Notional missing in latest Exposure row")

today = datetime.utcnow().strftime("%Y-%m-%d")

# Append a NEW row (daily snapshot)
new_row = last_row + 1
ws[f"A{new_row}"] = today
ws[f"B{new_row}"] = equity
ws[f"C{new_row}"] = notional

wb.save(EXCEL_PATH)

print(f"Exposure snapshot recorded | Equity={equity}, Notional={notional}")
