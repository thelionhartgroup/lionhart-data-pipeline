import requests
import openpyxl
from datetime import datetime
import os
import csv

# ================= CONFIG =================
EXCEL_PATH = "Macro_Data.xlsx"
CSV_PATH = "live_prices.csv"
SHEET_NAME = "Live_Prices"

ALPACA_API_KEY = os.getenv("ALPACA_API_KEY")
ALPACA_SECRET_KEY = os.getenv("ALPACA_SECRET_KEY")

print("DEBUG: ALPACA_API_KEY present =", bool(ALPACA_API_KEY))

if not ALPACA_API_KEY or not ALPACA_SECRET_KEY:
    raise RuntimeError("Alpaca API credentials are missing. Check GitHub Secrets.")

# ================= FETCH PRICE =================
def get_price(symbol: str) -> float:
    url = f"https://data.alpaca.markets/v2/stocks/{symbol}/quotes/latest"
    headers = {
        "APCA-API-KEY-ID": ALPACA_API_KEY,
        "APCA-API-SECRET-KEY": ALPACA_SECRET_KEY
    }

    response = requests.get(url, headers=headers, timeout=10)
    response.raise_for_status()

    data = response.json()
    price = data["quote"]["ap"]  # 'ap' is the ask price, adjust if you need bid or another field

    if price is None:
        raise ValueError(f"No price returned for symbol: {symbol}")

    return price

# ================= MAIN =================
def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    updated_symbols = 0

    # Update prices in Excel
    for row in range(2, ws.max_row + 1):
        symbol = ws[f"A{row}"].value

        if not symbol:
            continue

        try:
            price = get_price(symbol)
            ws[f"B{row}"] = price
            ws[f"C{row}"] = timestamp
            updated_symbols += 1
        except Exception as e:
            print(f"Error updating {symbol}: {e}")

    wb.save(EXCEL_PATH)

    print(f"Live prices updated successfully | Symbols updated: {updated_symbols}")

    # Export CSV (Mac-compatible, Power Query friendly)
    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Symbol", "Price", "Timestamp"])

        for row in range(2, ws.max_row + 1):
            writer.writerow([
                ws[f"A{row}"].value,
                ws[f"B{row}"].value,
                ws[f"C{row}"].value
            ])

# ================= ENTRY POINT =================
if __name__ == "__main__":
    main()
